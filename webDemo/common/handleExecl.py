from openpyxl import load_workbook
from singleton import Singleton

class handleExcelData(Singleton):
    def __init__(self, excel_path, name=None):
        self.execl_path = excel_path
        self.name = name
        wb = load_workbook(self.execl_path)
        if self.name is None:
            self.ws = wb.active

        else:
            self.ws = wb[self.name]
        self.head_date_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        # self.head_date_tuple=[i.value for i in self.ws[1]]

    def getExcelData(self):

        one_list = []
        for one_tuple in tuple(self.ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(self.head_date_tuple, one_tuple)))
        return one_list
        # rows=list(self.ws.rows)
        # datas=[]
        # for row in rows[1:]:
        #     data=[]
        #     for cell in row:
        #         data.append(cell.value)
        #         data_dict=dict(zip(self.head_date_tuple,data))
        #     datas.append(data_dict)
        # return datas


    def write_data(self, row, request):
        if isinstance(row, int) and (2 <= row <= self.ws.max_row):
            self.ws.cell(row, column=self.head_date_tuple.index("request") + 1, value=request)
        self.ws.save(self.execl_path)


if __name__ == '__main__':
    do_excelData = handleExcelData(excel_path="D:\demo\新建 XLSX 工作表.xlsx")
    do_excelData.getExcelData()
